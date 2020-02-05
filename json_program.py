# -*- coding: utf-8 -*-
"""
Created on Thu Jul 18 20:51:05 2019

@author: murmeg01
"""

import pandas as pd
import openpyxl
import json

path = "C:\\Users\\murmeg01\\Documents\\00 MBA EPBA-08\\Project\\Level 2\\Most_Viewed_Input.xlsx"

wb_obj = openpyxl.load_workbook(path) 
sheet_obj = wb_obj.active 
max_row = sheet_obj.max_row

df1 = pd.DataFrame()
df2 = pd.DataFrame()

def unpack(df, column, fillna=None):
    ret = None
    fillna=0
    if fillna is None:
        tmp = pd.DataFrame((d for idx, d in df[column].iteritems()))
        ret = pd.concat([df.drop(column,axis=1), tmp], axis=1)
    else:
        tmp = pd.DataFrame((d for idx, d in 
        df[column].iteritems())).fillna(fillna)
        ret = pd.concat([df.drop(column,axis=1), tmp], axis=1)
    return ret

for row in range(1,max_row+1):
    cell_obj = sheet_obj.cell(row = row, column = 1) 
    #print("Cell value :",cell_obj.value)
    json_data = json.loads(cell_obj.value)
    #print("Data value :",type(json_data))
    
    for key, value in json_data.items():
        #print(key)
        #print(len(value))
        #print("type :",type(json_data[key]))
        df1 = pd.DataFrame({'Channel_ID':key,
                            'Video_Data':value})
    
        df1 = df1.loc[~df1.index.duplicated(keep='first')]
        df1 = unpack(df1, 'Video_Data', 0)
        
        for i in range(0,2):
            df2 = df1.append(df2)

df2.to_excel("most_viewed_output.xlsx",index=False)